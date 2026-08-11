import io
import os
import streamlit as st
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from utils.translations import init_language, t

# ==============================================================================
# 0. Page Configuration & Language Init
# ==============================================================================
st.set_page_config(
    page_title="Export Cleaned Dataset & Reports",
    page_icon="📦",
    layout="wide"
)

# يقرأ اللغة المختارة ويظهر القائمة الجانبية
init_language()

# قراءة الـ CSS الموحد
try:
    with open("assets/style.css", encoding="utf-8") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
except FileNotFoundError:
    pass

st.title("📦 Export Project Data & Analytics")
st.write(t("sub_title") if t("sub_title") != "sub_title" else "Download your fully cleaned dataset as an Excel Table, CSV, or export project assets.")

st.divider()

# ==============================================================================
# 1. Helper Function: Convert DF to Styled Excel Table with Auto Column Width
# ==============================================================================
def convert_df_to_styled_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    
    # التأكد من عدم وجود أعمدة فارغة أو مكررة لمنع تلف ملف الإكسيل
    clean_export_df = df.copy()
    clean_export_df.columns = [str(col) if col is not None else "Unnamed" for col in clean_export_df.columns]

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        clean_export_df.to_excel(writer, index=False, sheet_name='Cleaned_Data')
        
        workbook = writer.book
        worksheet = writer.sheets['Cleaned_Data']
        
        max_row = len(clean_export_df) + 1
        max_col = len(clean_export_df.columns)
        
        if max_row > 1 and max_col > 0:
            end_col_letter = get_column_letter(max_col)
            tab_range = f"A1:{end_col_letter}{max_row}"
            
            # إضافة تنسيق جدول إكسيل الأزرق مع اسم فريد
            tab = Table(displayName="CleanedDataTableFit", ref=tab_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            
            # 📐 ضبط عرض الأعمدة تلقائياً بناءً على محتوى النص
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                # تحديد عرض العمود مع هامش بسيط وتحديد حد أقصى للحماية (50)
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)
        
    return output.getvalue()

# ==============================================================================
# 2. Smart Dataset Detection
# ==============================================================================
export_df = None

possible_keys = ['cleaned_df', 'df', 'raw_df', 'data', 'dataset', 'uploaded_df']

for key in possible_keys:
    if key in st.session_state and st.session_state[key] is not None:
        if isinstance(st.session_state[key], pd.DataFrame):
            export_df = st.session_state[key]
            st.success("✅ Active dataset detected automatically (from session context)!")
            break

# ==============================================================================
# 3. Render Export Buttons
# ==============================================================================
if export_df is not None:
    st.subheader("📊 Download Processed Dataset")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 🟢 Excel Spreadsheet (`.xlsx`)")
        st.info("Includes built-in **Excel Table formatting** with auto-adjusted column widths.")
        
        excel_bytes = convert_df_to_styled_excel(export_df)
        
        st.download_button(
            label="📥 Download Excel Table (.xlsx)",
            data=excel_bytes,
            file_name="DataPilot_Cleaned_Dataset.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col2:
        st.markdown("### 🔵 Raw CSV File (`.csv`)")
        st.info("Standard comma-separated text file suitable for database integration.")
        
        # التعديل لترميز utf-8-sig لضمان عرض الحروف العربية بشكل متوافق مع Excel
        csv_data = export_df.to_csv(index=False).encode('utf-8-sig')
        
        st.download_button(
            label="📥 Download CSV Dataset (.csv)",
            data=csv_data,
            file_name="DataPilot_Cleaned_Dataset.csv",
            mime="text/csv",
            use_container_width=True
        )

    st.divider()

    st.subheader("👁️ Dataset Quick Summary")
    st.write(f"Total Rows: **{export_df.shape[0]:,}** | Total Columns: **{export_df.shape[1]}**")
    st.dataframe(export_df.head(10), use_container_width=True)

else:
    st.error(t("no_dataset") if t("no_dataset") != "no_dataset" else "❌ No active dataset found in memory.")